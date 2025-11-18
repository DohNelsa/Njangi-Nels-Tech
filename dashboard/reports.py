"""
Report generation utilities for exports
"""
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import redirect
from django.db.models import Sum, Count
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from contributions.models import Contribution, Withdrawal, TransactionLog
from loans.models import Loan
from members.models import Member
from meetings.models import Meeting


@login_required
def export_contributions_report(request):
    """Export contributions to Excel"""
    # Check if user is admin
    if not request.user.is_staff and not (hasattr(request.user, 'member_profile') and request.user.member_profile.is_admin()):
        messages.error(request, 'Only administrators can export reports.')
        return redirect('dashboard:index')
    
    contributions = Contribution.objects.all().order_by('-date')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Contributions Report"
    
    # Headers
    headers = ['Date', 'Member', 'Amount', 'Description', 'Recorded By', 'Recorded At']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for contribution in contributions:
        ws.append([
            contribution.date,
            contribution.member.name,
            contribution.amount,
            contribution.description,
            contribution.created_by.username if contribution.created_by else '',
            contribution.created_at.strftime('%Y-%m-%d %H:%M')
        ])
    
    # Total row
    total = contributions.aggregate(total=Sum('amount'))['total'] or 0
    ws.append(['', 'TOTAL', total, '', '', ''])
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=contributions_report.xlsx'
    wb.save(response)
    return response


@login_required
def export_members_report(request):
    """Export members to Excel"""
    # Check if user is admin
    if not request.user.is_staff and not (hasattr(request.user, 'member_profile') and request.user.member_profile.is_admin()):
        messages.error(request, 'Only administrators can export reports.')
        return redirect('dashboard:index')
    
    members = Member.objects.all().order_by('name')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Members Report"
    
    # Headers
    headers = ['Name', 'Phone', 'Email', 'Role', 'Date Joined', 'Total Contributions', 'Current Balance', 'Status']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for member in members:
        ws.append([
            member.name,
            member.phone,
            member.email,
            member.get_role_display(),
            member.date_joined.strftime('%Y-%m-%d'),
            member.get_total_contributions(),
            member.get_current_balance(),
            'Active' if member.is_active else 'Inactive'
        ])
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=members_report.xlsx'
    wb.save(response)
    return response


@login_required
def export_transaction_logs(request):
    """Export transaction logs to Excel"""
    # Check if user is admin
    if not request.user.is_staff and not (hasattr(request.user, 'member_profile') and request.user.member_profile.is_admin()):
        messages.error(request, 'Only administrators can export reports.')
        return redirect('dashboard:index')
    
    logs = TransactionLog.objects.all().order_by('-created_at')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaction Logs"
    
    # Headers
    headers = ['Date', 'Type', 'Member', 'Amount', 'Description', 'Recorded By']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for log in logs:
        ws.append([
            log.created_at.strftime('%Y-%m-%d %H:%M'),
            log.get_transaction_type_display(),
            log.member.name if log.member else '',
            log.amount,
            log.description,
            log.created_by.username if log.created_by else ''
        ])
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=transaction_logs.xlsx'
    wb.save(response)
    return response

